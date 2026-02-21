import sqlite3
import tkinter as tk
import pyperclip
import math
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from datetime import datetime
from tkcalendar import Calendar
from tkinter import ttk
from datetime import datetime, timedelta
import tkinter.font as tkFont
from tkinter import PhotoImage
import os
from PIL import Image, ImageTk
from tkinter import font
from datetime import datetime



# Global variable to count imported Excel files
import_count = 0

default_db_path = "E:/25 BIRBA/01-DataBase/Production_Database_Birba.db"
pic_dir = "E:/25 BIRBA/06-Hou Program Hub/code pic"


def connect_to_database(db_path=default_db_path):
    conn = sqlite3.connect(db_path)
    return conn

def format_percentage(value):
    if isinstance(value, (int, float)):
        return f"{value * 100:.2f}%"
    else:
        return "0.00%"

def format_percentage_no_sign(value):
    if isinstance(value, (int, float)):
        return f"{value * 100:.2f}"
    else:
        return "0.00%"

# Function to format date values
def format_date(date_str):
    if isinstance(date_str, datetime):
        return date_str.strftime('%d-%b-%Y')
    else:
        return "01-Jan-1990"

def convert_to_julian_day(date_str):
    date_obj = datetime.strptime(date_str, '%d-%b-%Y')
    julian_day = date_obj.toordinal() + 1721424
    return int(julian_day)  # Convert to integer

# Function to extract data from Excel file
def extract_data_from_excel(file_path):
    try:
        wb = load_workbook(file_path, data_only=True)
        sheets = wb.sheetnames
        # Filter sheets that start with month names
        month_sheets = [sheet for sheet in sheets if
                        sheet[:3] in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov',
                                      'Dec']]

        data = {}

        for sheet_name in month_sheets:
            ws = wb[sheet_name]
            # Example: Process data for each sheet
            data[sheet_name] = {
            'Daily_Nodes': {
                'Date': format_date(ws['B1'].value) if ws['B1'].value else "01-January-1990",
                'JulianDay': convert_to_julian_day(format_date(ws['B1'].value)) if ws['B1'].value else convert_to_julian_day("01-January-1990"),
                'Nodes_Inova_Layout': ws['Q48'].value if ws['Q48'].value else 0,
                'Nodes_Inova_Pickup': ws['Q49'].value if ws['Q49'].value else 0,
                'Nodes_GTI_layout': ws['Q50'].value if ws['Q50'].value else 0,
                'Nodes_GTI_Pickup': ws['Q51'].value if ws['Q51'].value else 0,
                'Front_Crew_Inova_Teams': ws['T48'].value if ws['T48'].value else 0,
                'Back_Crew_Inova_Teams': ws['T49'].value if ws['T49'].value else 0,
                'Front_Crew_GTI_Teams': ws['T50'].value if ws['T50'].value else 0,
                'Back_Crew_GTI_Teams': ws['T51'].value if ws['T51'].value else 0,
                'Drone_Percentage': format_percentage(ws['Q45'].value),
                'TS_HQ_other_Teams': ws['T47'].value if ws['T47'].value else 0
            },
            'Daily_Recording': {
                'Date': format_date(ws['B1'].value) if ws['B1'].value else "01-January-1990",
                'JulianDay': convert_to_julian_day(format_date(ws['B1'].value)) if ws['B1'].value else convert_to_julian_day("01-January-1990"),
                'Recording_Hours': round(ws['B26'].value, 3) if isinstance(ws['B26'].value, (int, float)) else 0,
                'Avg_VPS_per_rec_hrs': round(ws['B15'].value, 0) if isinstance(ws['B15'].value, (int, float)) else 0,
                'Max_VPS_hr': ws['B16'].value if ws['B16'].value else 0,
                'APP_over_CTM': format_percentage(ws['B19'].value),
                'Skips': ws['F13'].value if ws['F13'].value else 0,
                'Daily_TCF': round(ws['B14'].value, 3) if isinstance(ws['B14'].value, (int, float)) else 0,
            },
            'Daily_VPs': {
                'Date': format_date(ws['B1'].value) if ws['B1'].value else "01-January-1990",
                'JulianDay': convert_to_julian_day(format_date(ws['B1'].value)) if ws['B1'].value else convert_to_julian_day("01-January-1990"),
                'Flat': ws['N7'].value if ws['N7'].value else 0,
                'Flat_Percentage': format_percentage(ws['O7'].value),
                'Rough': ws['N8'].value if ws['N8'].value else 0,
                'Rough_Percentage': format_percentage(ws['O8'].value),
                'Facilities': ws['N9'].value if ws['N9'].value else 0,
                'Facilities_Percentage': format_percentage(ws['O9'].value),
                'Sand_Dunes': ws['N10'].value if ws['N10'].value else 0,
                'Sand_Dunes_Percentage': format_percentage(ws['O10'].value),
                'Sabkha': ws['N11'].value if ws['N11'].value else 0,
                'Sabkha_Percentage': format_percentage(ws['O11'].value),
                'Total_VPs_APP': ws['N13'].value if ws['N13'].value else 0,
                'CTM': int(float(ws['B17'].value)) if ws['B17'].value else 0,
                'Cumulative_VPs': ws['B22'].value if ws['B22'].value else 0,
                'Comments': f"{ws['P1'].value}\n{ws['P2'].value}\n{ws['P3'].value}\n{ws['P4'].value}\n{ws['P5'].value}\n{ws['P6'].value}\n{ws['P7'].value}"
            },
            'Downtime': {
                'Date': format_date(ws['B1'].value) if ws['B1'].value else "01-January-1990",
                'JulianDay': convert_to_julian_day(format_date(ws['B1'].value)) if ws['B1'].value else convert_to_julian_day("01-January-1990"),
                'Recording_Equipment_Faults': round(ws['B39'].value, 3) if ws['B39'].value else 0,
                'Vibrator_Faults': round(ws['B40'].value, 3) if ws['B40'].value else 0,
                'Incidents': round(ws['B41'].value, 3) if ws['B41'].value else 0,
                'Legal_Action_Labor_Dispute': round(ws['B42'].value, 3) if ws['B42'].value else 0,
                'Company_Instructions': round(ws['B43'].value, 3) if ws['B43'].value else 0,
                'Contractor_Generated_Noise': round(ws['B44'].value, 3) if ws['B44'].value else 0,
                'Other_DT': round(ws['B45'].value, 3) if ws['B45'].value else 0,
            },
            'Operational_Time': {
                'Date': format_date(ws['B1'].value) if ws['B1'].value else "01-January-1990",
                'JulianDay': convert_to_julian_day(format_date(ws['B1'].value)) if ws['B1'].value else convert_to_julian_day("01-January-1990"),
                'Recording_Hours': round(ws['B26'].value, 3) if isinstance(ws['B26'].value, (int, float)) else 0,
                'VCU_move_up': round(ws['B27'].value, 3) if isinstance(ws['B27'].value, (int, float)) else 0,
                'Daily_Monthly_Testing': round(ws['B28'].value, 3) if isinstance(ws['B28'].value, (int, float)) else 0,
                'Waiting_source_Layout_Shift_change': round(ws['B29'].value, 3) if isinstance(ws['B29'].value, (int, float)) else 0,
            },
            'Stand_By_Time': {
                'Date': format_date(ws['B1'].value) if ws['B1'].value else "01-January-1990",
                'JulianDay': convert_to_julian_day(format_date(ws['B1'].value)) if ws['B1'].value else convert_to_julian_day("01-January-1990"),
                'Company_Suspension_Awaiting_Company': round(ws['B34'].value, 3) if ws['B34'].value else 0,
                'Company_Requested_Tests': round(ws['B35'].value, 3) if ws['B35'].value else 0,
                'Beyond_Contractor_Control': round(ws['B36'].value, 3) if ws['B36'].value else 0,
                'Camp_Move': round(ws['B37'].value, 3) if ws['B37'].value else 0,
            },
            'Total_Time': {
                'Date': format_date(ws['B1'].value) if ws['B1'].value else "01-January-1990",
                'JulianDay': convert_to_julian_day(format_date(ws['B1'].value)) if ws['B1'].value else convert_to_julian_day("01-January-1990"),
                'Total_Operation_Time': round(
                    (ws['B26'].value if ws['B26'].value else 0) +
                    (ws['B27'].value if ws['B27'].value else 0) +
                    (ws['B28'].value if ws['B28'].value else 0) +
                    (ws['B30'].value if ws['B30'].value else 0) +
                    (ws['B31'].value if ws['B31'].value else 0) +
                    (ws['B32'].value if ws['B32'].value else 0),
                    3),
                'Total_Downtime': round(
                    (ws['B39'].value if ws['B39'].value else 0) +
                    (ws['B40'].value if ws['B40'].value else 0) +
                    (ws['B41'].value if ws['B41'].value else 0) +
                    (ws['B42'].value if ws['B42'].value else 0) +
                    (ws['B43'].value if ws['B43'].value else 0) +
                    (ws['B44'].value if ws['B44'].value else 0) +
                    (ws['B45'].value if ws['B45'].value else 0),
                    3),
                'Total_Stand_By_Time': round(
                    (ws['B34'].value if ws['B34'].value else 0) +
                    (ws['B35'].value if ws['B35'].value else 0) +
                    (ws['B36'].value if ws['B36'].value else 0) +
                    (ws['B37'].value if ws['B37'].value else 0),
                    3),
            },
                'Daily_HSE_Statistics': {
                    'Date': format_date(ws['B1'].value) if ws['B1'].value else "01-January-1990",
                    'JulianDay': convert_to_julian_day(format_date(ws['B1'].value)) if ws['B1'].value else convert_to_julian_day("01-January-1990"),
                    'Stop_cards': ws['Q35'].value if ws['Q35'].value else 0,
                    'LTI': ws['Q36'].value if ws['Q36'].value else 0,
                    'FAC': ws['Q37'].value if ws['Q37'].value else 0,
                    'MTC': ws['Q38'].value if ws['Q38'].value else 0,
                    'RWC': ws['Q39'].value if ws['Q39'].value else 0,
                    'Oil_spill': ws['T34'].value if ws['T34'].value else 0,
                    'Incident': ws['T35'].value if ws['T35'].value else 0,
                    'Near_Miss': ws['U35'].value if ws['U35'].value else 0,
                    'Medevac': ws['T36'].value if ws['T36'].value else 0,
                    'Drills': ws['T37'].value if ws['T37'].value else 0,
                    'Inspections_Audits': ws['T38'].value if ws['T38'].value else 0,
                    'LSR_Violation': ws['T39'].value if ws['T39'].value else 0,
                },
            'Weather': {
                'Date': format_date(ws['B1'].value) if ws['B1'].value else "01-January-1990",
                'JulianDay': convert_to_julian_day(format_date(ws['B1'].value)) if ws['B1'].value else convert_to_julian_day("01-January-1990"),
                'Conditions': ws['P15'].value if ws['P15'].value else "Sunny",
                'Rain': ws['R15'].value if ws['R15'].value else "N",
                'Max_Temp': ws['S15'].value if ws['S15'].value else 0,
                'Min_Temp': ws['T15'].value if ws['T15'].value else 0
            },
            'Additional_Info': {
                'Date': format_date(ws['B1'].value) if ws['B1'].value else "01-January-1990",
                'JulianDay': convert_to_julian_day(format_date(ws['B1'].value)) if ws['B1'].value else convert_to_julian_day("01-January-1990"),
                'No_Percentage_APP_over_CTM': format_percentage_no_sign(ws['B19'].value),
                'Total_VPs_APP': ws['N13'].value if ws['N13'].value else 0
            }
        }

        return data

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while reading Excel file '{file_path}': {str(e)}")
        return None


def import_data():
    global import_count, db_path, file_paths

    db_path = database_entry.get()
    file_paths = file_entry.get().split(';')

    if not db_path or not file_paths:
        messagebox.showerror("Error", "Please select both a database and at least one file.")
        return

    import_count = 0  # Reset import counter

    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        for file_path in file_paths:
            data = extract_data_from_excel(file_path)

            if data:
                for sheet_name, sheet_data in data.items():
                    date = sheet_data['Daily_Nodes']['Date']
                    for table_name in sheet_data.keys():
                        cursor.execute(f"DELETE FROM {table_name} WHERE Date = ?", (date,))

                    for table_name, table_data in sheet_data.items():
                        columns = ', '.join(table_data.keys())
                        placeholders = ', '.join(['?'] * len(table_data))
                        sql = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
                        cursor.execute(sql, list(table_data.values()))

                        # Calculate AVG_Production after inserting or updating Additional_Info
                        if table_name == 'Additional_Info':
                            calculate_avg_production(conn, cursor)

                # Commit changes to database after each file import
                conn.commit()
                import_count += 1

            else:
                messagebox.showerror("Error", f"Failed to extract data from Excel, Check Date & sheet name in '{file_path}'.")

        if import_count > 0:
            messagebox.showinfo("Success", f"{import_count} Daily Report(s) imported successfully.")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while importing data: {str(e)}")
        conn.rollback()

    finally:
        conn.close()

def calculate_avg_production(conn, cursor):
    try:
        cursor.execute("""
            UPDATE Additional_Info AS ai
            SET AVG_Production = (
                SELECT IFNULL(CAST(AVG(ai2.Total_VPs_APP) AS INTEGER), 0)
                FROM Additional_Info AS ai2
                WHERE ai2.JulianDay <= ai.JulianDay 
                  AND ai2.Total_VPs_APP <> 0
            )
        """)

        conn.commit()

    except Exception as e:
        conn.rollback()
        raise e


def display_info():
    date = date_entry.get()

    if not date:
        messagebox.showerror("Error", "Please enter a date.")
        return

    db_path = database_entry.get() or default_db_path  # Use the default path if no path is specified

    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Retrieve and display Weather Conditions
        cursor.execute("SELECT Conditions FROM Weather WHERE Date = ?", (date,))
        conditions = cursor.fetchone()
        conditions_value.config(text=conditions[0] if conditions else "N/A")

        # Retrieve Production
        cursor.execute("SELECT Total_VPs_APP FROM Daily_VPs WHERE Date = ?", (date,))
        Total_VPs_APP = cursor.fetchone()
        Total_VPs_APP_value.config(text=Total_VPs_APP[0] if Total_VPs_APP else "N/A")

        cursor.execute("SELECT Flat FROM Daily_VPs WHERE Date = ?", (date,))
        Flat = cursor.fetchone()
        Flat_value.config(text=Flat[0] if Flat else "N/A")

        cursor.execute("SELECT Rough FROM Daily_VPs WHERE Date = ?", (date,))
        Rough = cursor.fetchone()
        Rough_value.config(text=Rough[0] if Rough else "N/A")

        cursor.execute("SELECT Facilities FROM Daily_VPs WHERE Date = ?", (date,))
        Facilities = cursor.fetchone()
        Facilities_value.config(text=Facilities[0] if Facilities else "N/A")

        cursor.execute("SELECT Sand_Dunes FROM Daily_VPs WHERE Date = ?", (date,))
        Sand_Dunes = cursor.fetchone()
        Sand_Dunes_value.config(text=Sand_Dunes[0] if Sand_Dunes else "N/A")

        cursor.execute("SELECT Sabkha FROM Daily_VPs WHERE Date = ?", (date,))
        Sabkha = cursor.fetchone()
        Sabkha_value.config(text=Sabkha[0] if Sabkha else "N/A")


        # Retrieve and display Total_Operation_Time
        cursor.execute("SELECT Total_Operation_Time FROM Total_Time WHERE Date = ?", (date,))
        Total_Operation_Time = cursor.fetchone()
        Recording_Hours_value.config(text=Total_Operation_Time[0] if Total_Operation_Time else "N/A")

        # Retrieve and display Total_Downtime
        cursor.execute("SELECT Total_Downtime FROM Total_Time WHERE Date = ?", (date,))
        DR_Down_time = cursor.fetchone()
        DR_Down_time_value.config(text=DR_Down_time[0] if DR_Down_time else "N/A")

        # Retrieve and display Total_Stand_By_Time
        cursor.execute("SELECT Total_Stand_By_Time FROM Total_Time WHERE Date = ?", (date,))
        Total_Stand_By_Time = cursor.fetchone()
        Total_Stand_By_Time_value.config(text=Total_Stand_By_Time[0] if Total_Stand_By_Time else "N/A")

        # Retrieve and display CTM
        cursor.execute("SELECT CTM FROM Daily_VPs WHERE Date = ?", (date,))
        CTM = cursor.fetchone()
        CTM_value.config(text=CTM[0] if CTM else "N/A")

        # Retrieve and display APP_over_CTM
        cursor.execute("SELECT APP_over_CTM FROM Daily_Recording WHERE Date = ?", (date,))
        APP_over_CTM = cursor.fetchone()
        APP_over_CTM_value.config(text=APP_over_CTM[0] if APP_over_CTM else "N/A")

        cursor.execute("SELECT Final_APP_CTM FROM Additional_Info WHERE Date = ?", (date,))
        Final_APP_CTM = cursor.fetchone()
        Final_APP_CTM_value.config(
            text=f"{float(Final_APP_CTM[0].replace('%', '')):.2f}%" if Final_APP_CTM else "N/A"
        )

        cursor.execute("SELECT AVG_Production FROM Additional_Info WHERE Date = ?", (date,))
        AVG_Production = cursor.fetchone()
        AVG_Production_value.config(text=AVG_Production[0] if AVG_Production else "N/A")

        # Retrieve and display Stop_cards from Daily_HSE_Statistics
        cursor.execute("SELECT Stop_cards FROM Daily_HSE_Statistics WHERE Date = ?", (date,))
        Stop_cards = cursor.fetchone()
        Stop_cards_value.config(text=Stop_cards[0] if Stop_cards else "N/A")

        cursor.execute("SELECT Inspections_Audits FROM Daily_HSE_Statistics WHERE Date = ?", (date,))
        Inspections_Audits = cursor.fetchone()
        Inspections_Audits_value.config(text=Inspections_Audits[0] if Inspections_Audits else "N/A")

        cursor.execute("SELECT Drills FROM Daily_HSE_Statistics WHERE Date = ?", (date,))
        Drills = cursor.fetchone()
        Drills_value.config(text=Drills[0] if Drills else "N/A")

        cursor.execute("SELECT Comments FROM Daily_VPs WHERE Date = ?", (date,))
        Comments = cursor.fetchone()
        if Comments and Comments[0]:
            Comments_value.delete("1.0", tk.END)  # Clear previous text
            Comments_value.insert(tk.END, Comments[0])
        else:
            Comments_value.delete("1.0", tk.END)
            Comments_value.insert(tk.END, "N/A")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while displaying data: {str(e)}")


def select_database():
    database_path = filedialog.askopenfilename(filetypes=[("SQLite Database", "*.db")])
    if database_path:
        database_entry.delete(0, tk.END)
        database_entry.insert(0, database_path)


def select_file():
    file_paths = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx; *.xls")])
    if file_paths:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, ';'.join(file_paths))

def pick_date():
    def print_sel():
        selected_date = cal.selection_get()
        formatted_date = selected_date.strftime("%d-%b-%Y")
        date_entry.delete(0, tk.END)  # Clear any existing text
        date_entry.insert(0, formatted_date)
        top.destroy()

    top = tk.Toplevel(root)
    cal = Calendar(top, selectmode='day')
    cal.grid(row=0, column=0, padx=10, pady=10)

    ttk.Button(top, text="Ok", command=print_sel).grid(row=1, column=0, pady=20)

# Calculate yesterday's date
yesterday = datetime.now() - timedelta(1)
formatted_yesterday = yesterday.strftime("%d-%b-%Y")

# Define the font attributes: bold and underline
font_attributes = ("Helvetica", 12, "bold", "underline")
font_attributes1 = ("Helvetica", 12, "bold")


# Function to copy the displayed information to the clipboard
def copy_to_clipboard():
    clipboard_text = (
        f"Weather: {conditions_value.cget('text')}\n"
        f"Production: {Total_VPs_APP_value.cget('text')}VPs "
        f"(Terrain Flat:{Flat_value.cget('text')}VPs, "
        f"Rough:{Rough_value.cget('text')}VPs, "
        f"Facilities:{Facilities_value.cget('text')}VPs, "
        f"Sand Dunes:{Sand_Dunes_value.cget('text')}VPs)\n"
        f"Production time: {Recording_Hours_value.cget('text')}Hrs\n"
        f"Downtime: {DR_Down_time_value.cget('text')}Hrs\n"
        f"Standby time: {Total_Stand_By_Time_value.cget('text')}Hrs\n"
        f"CTM: {CTM_value.cget('text')}\n"
        f"APP/CTM: {APP_over_CTM_value.cget('text')}\n"
        f"Final APP/CTM: {Final_APP_CTM_value.cget('text')}\n"
        f"Average production of this project: {AVG_Production_value.cget('text')}\n"
        f"Daily HSE Statistics: Stop cards: {Stop_cards_value.cget('text')}, "
        f"Audits: {Inspections_Audits_value.cget('text')}, "
        f"Drills: {Drills_value.cget('text')}."
    )
    root.clipboard_clear()
    root.clipboard_append(clipboard_text)

# GUI setup using tkinter
root = tk.Tk()
root.title("Daily Report Import")

# Create a custom font for the project name with underline, bold, and italic
project_font = font.Font(family='Helvetica', size=24, weight='bold', slant='italic', underline=True)

# Display the project name with custom font and color
project_name_label = tk.Label(root, text="Birba", font=project_font, fg='silver')
project_name_label.grid(row=1, column=3, columnspan=11, padx=5, pady=20)


database_label = tk.Label(root, text="Select SQLite Database:")
database_label.grid(row=0, column=0, padx=5, pady=5)
database_entry = tk.Entry(root, width=50)
database_entry.grid(row=0, column=1, padx=5, pady=5)
database_entry.insert(0, default_db_path)
database_button = tk.Button(root, text="Browse", command=select_database)
database_button.grid(row=0, column=2, padx=5, pady=5)

file_label = tk.Label(root, text="Select Daily Report(s):")
file_label.grid(row=1, column=0, padx=5, pady=5)
file_entry = tk.Entry(root, width=50)
file_entry.grid(row=1, column=1, padx=5, pady=5)
file_button = tk.Button(root, text="Browse", command=select_file)
file_button.grid(row=1, column=2, padx=5, pady=5)

import_button = tk.Button(root, text="Import To Database", command=import_data)
import_button.grid(row=3, column=1, padx=5, pady=5)

bold_font = tkFont.Font(root, weight="bold")

date_label = tk.Label(root, text="Select Date:", font=bold_font)
date_label.grid(row=4, column=0, padx=10, pady=10)
date_entry = tk.Entry(root, width=12,font=bold_font, justify='center')
date_entry.grid(row=4, column=1, padx=5, pady=5)
date_entry.insert(0, formatted_yesterday)  # Set default date to yesterday


calender_image_path = os.path.join(pic_dir, "calender_icon.png")
vib_image_path = os.path.join(pic_dir, "vib_icon.jpg")
email_image_path = os.path.join(pic_dir, "email_icon.png")

# Open the image with PIL
image1 = Image.open(calender_image_path)
calender_resized_image = image1.resize((25, 25), Image.LANCZOS)

image2 = Image.open(vib_image_path)
vib_resized_image = image2.resize((75, 50), Image.LANCZOS)

image3 = Image.open(email_image_path)
email_resized_image = image3.resize((35, 25), Image.LANCZOS)

icon1 = ImageTk.PhotoImage(calender_resized_image)
icon2 = ImageTk.PhotoImage(vib_resized_image)
icon3 = ImageTk.PhotoImage(email_resized_image)

icon1_label = tk.Label(root, image=icon1)
icon1_label.grid(row=4, column=1, padx=(10, 0), pady=20, sticky="e")

icon2_label = tk.Label(root, image=icon2)
icon2_label.grid(row=4, column=4, padx=(10, 0), pady=20, sticky="e")

icon3_label = tk.Label(root, image=icon3)
icon3_label.grid(row=4, column=6, padx=(10, 0), pady=20, sticky="e")

icon1_label.image = icon1
icon2_label.image = icon2
icon3_label.image = icon3


ttk.Button(root, text="Choose Date", command=pick_date).grid(row=4, column=2, pady=20)

ttk.Button(root, text="Display Info", command=display_info).grid(row=4, column=5, pady=20)

ttk.Button(root, text="Copy Info", command=copy_to_clipboard).grid(row=4, column=7, pady=20)

# Labels to display imported information

conditions_value_label = tk.Label(root, text="Weather:", font=font_attributes)
conditions_value_label.grid(row=5, column=0, padx=5, pady=5)
conditions_value = tk.Label(root, text="")
conditions_value.grid(row=5, column=1, padx=5, pady=5)

Total_VPs_APP_label = tk.Label(root, text="Production:", font=font_attributes)
Total_VPs_APP_label.grid(row=6, column=0, padx=5, pady=5)
Total_VPs_APP_value = tk.Label(root, text="")
Total_VPs_APP_value.grid(row=6, column=1, padx=5, pady=5)

Flat_label = tk.Label(root, text="(Terrain Flat:")
Flat_label.grid(row=6, column=2, padx=5, pady=5)
Flat_value = tk.Label(root, text="")
Flat_value.grid(row=6, column=3, padx=5, pady=5)

Rough_label = tk.Label(root, text="Rough:")
Rough_label.grid(row=6, column=4, padx=5, pady=5)
Rough_value = tk.Label(root, text="")
Rough_value.grid(row=6, column=5, padx=5, pady=5)

Facilities_label = tk.Label(root, text="Facilities:")
Facilities_label.grid(row=6, column=6, padx=5, pady=5)
Facilities_value = tk.Label(root, text="")
Facilities_value.grid(row=6, column=7, padx=5, pady=5)

Sand_Dunes_label = tk.Label(root, text="Sand Dunes:")
Sand_Dunes_label.grid(row=6, column=8, padx=5, pady=5)
Sand_Dunes_value = tk.Label(root, text="")
Sand_Dunes_value.grid(row=6, column=9, padx=5, pady=5)

Sabkha_label = tk.Label(root, text="Sabkha:")
Sabkha_label.grid(row=6, column=10, padx=5, pady=5)
Sabkha_value = tk.Label(root, text="")
Sabkha_value.grid(row=6, column=11, padx=5, pady=5)

Last_va_label = tk.Label(root, text=")")
Last_va_label.grid(row=6, column=13, padx=5, pady=5)

Recording_Hours_label = tk.Label(root, text="Production time:", font=font_attributes)
Recording_Hours_label.grid(row=7, column=0, padx=5, pady=5)
Recording_Hours_value = tk.Label(root, text="")
Recording_Hours_value.grid(row=7, column=1, padx=5, pady=5)

DR_Down_time_label = tk.Label(root, text="Downtime:", font=font_attributes)
DR_Down_time_label.grid(row=8, column=0, padx=5, pady=5)
DR_Down_time_value = tk.Label(root, text="")
DR_Down_time_value.grid(row=8, column=1, padx=5, pady=5)

Total_Stand_By_Time_label = tk.Label(root, text="Standby time:", font=font_attributes)
Total_Stand_By_Time_label.grid(row=9, column=0, padx=5, pady=5)
Total_Stand_By_Time_value = tk.Label(root, text="")
Total_Stand_By_Time_value.grid(row=9, column=1, padx=5, pady=5)

CTM_label = tk.Label(root, text="CTM=", font=font_attributes)
CTM_label.grid(row=10, column=0, padx=5, pady=5)
CTM_value = tk.Label(root, text="")
CTM_value.grid(row=10, column=1, padx=5, pady=5)

APP_over_CTM_label = tk.Label(root, text="APP/CTM=", font=font_attributes)
APP_over_CTM_label.grid(row=11, column=0, padx=5, pady=5)
APP_over_CTM_value = tk.Label(root, text="")
APP_over_CTM_value.grid(row=11, column=1, padx=5, pady=5)

Final_APP_CTM_label = tk.Label(root, text="Final APP/CTM=", font=font_attributes)
Final_APP_CTM_label.grid(row=12, column=0, padx=5, pady=5)
Final_APP_CTM_value = tk.Label(root, text="")
Final_APP_CTM_value.grid(row=12, column=1, padx=5, pady=5)

AVG_Production_label = tk.Label(root, text="Average production of this project:", font=font_attributes)
AVG_Production_label.grid(row=13, column=0, padx=5, pady=5)
AVG_Production_value = tk.Label(root, text="")
AVG_Production_value.grid(row=13, column=1, padx=5, pady=5)

HSE_label = tk.Label(root, text="Daily HSE Statistics:", font=font_attributes)
HSE_label.grid(row=14, column=0, padx=5, pady=5)

Stop_cards_label = tk.Label(root, text="Stop cards:")
Stop_cards_label.grid(row=14, column=1, padx=5, pady=5)
Stop_cards_value = tk.Label(root, text="")
Stop_cards_value.grid(row=14, column=2, padx=5, pady=5)

Inspections_Audits_label = tk.Label(root, text="Audits:")
Inspections_Audits_label.grid(row=14, column=3, padx=5, pady=5)
Inspections_Audits_value = tk.Label(root, text="")
Inspections_Audits_value.grid(row=14, column=4, padx=5, pady=5)

Drills_label = tk.Label(root, text="Drills:")
Drills_label.grid(row=14, column=5, padx=5, pady=5)
Drills_value = tk.Label(root, text="")
Drills_value.grid(row=14, column=6, padx=5, pady=5)

comments_frame = tk.Frame(root)
comments_frame.grid(row=20, column=0, columnspan=10, pady=10, sticky="w")
Comments_label = tk.Label(comments_frame, text="Comments:", font=font_attributes1, bg="white", relief="ridge")
Comments_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
Comments_value = tk.Text(comments_frame, height=5, width=120, wrap="none")
Comments_value.grid(row=0, column=1, padx=5, pady=5)

bottom_space = tk.Frame(root, height=70, bg="white")
bottom_space.grid(row=21, column=0, columnspan=10)

# ---------------------------------------------
# Signature
# ---------------------------------------------
tk.Label(root, text="Hussein Al Shibli  |  BGP Oman APC", anchor="w").place(x=20, y=740)
# ---------------------------------------------

root.mainloop()